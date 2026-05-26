from __future__ import annotations

from dataclasses import dataclass, replace
import re
import shutil
import tempfile
import uuid
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidocr_onnxruntime import RapidOCR

CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ETC_NS = "http://www.wps.cn/officeDocument/2017/etCustomData"
CELL_IMAGE_REL_TYPE = "http://www.wps.cn/officeDocument/2020/cellImage"
IMAGE_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
CELL_IMAGE_CONTENT_TYPE = "application/vnd.wps-officedocument.cellimage+xml"
DEFAULT_ROW_HEIGHT = 73.05
IMAGE_EXT_CX = "4848225"
IMAGE_EXT_CY = "7200900"
HEADER_LABELS = [
    "交易单号",
    "商户单号",
    "交易金额",
    "交易时间",
    "咨询时间",
    "问题描述",
    "退款情况",
    "投诉时间",
    "处理时间",
    "处理结果",
    "订单商品说明",
    "未退款情况说明",
]
RECORD_HEADER = "沟通记录"
BASE_OUTPUT_NAME = "处理表格.xlsx"

ET.register_namespace("", SHEET_NS)
ET.register_namespace("a", A_NS)
ET.register_namespace("etc", ETC_NS)
ET.register_namespace("r", R_NS)
ET.register_namespace("xdr", XDR_NS)


@dataclass(slots=True)
class ExportComplaintData:
    transaction_id: str = ""
    merchant_order_id: str = ""
    transaction_amount: str = ""
    transaction_time: str = ""
    inquiry_time: str = ""
    problem_description: str = ""
    refund_summary: str = "未退款"
    refund_status: str = "未退款"
    complaint_dir: Path | None = None


@dataclass(slots=True)
class _CellImageEntry:
    cell_ref: str
    display_id: str
    rel_id: str
    media_name: str
    description: str


def find_latest_run_dir(output_root: Path) -> Path:
    root = Path(output_root)
    if not root.exists():
        raise FileNotFoundError(f"采集输出目录不存在：{root}")

    candidates = [
        path
        for path in root.iterdir()
        if path.is_dir() and re.fullmatch(r"\d{8}_\d{6}", path.name)
    ]
    if not candidates:
        raise FileNotFoundError(f"未找到采集目录：{root}")
    return max(candidates, key=lambda item: item.name)


def list_record_images_for_excel(complaint_dir: Path) -> list[Path]:
    images = sorted(
        [path for path in complaint_dir.iterdir() if path.is_file() and path.suffix.lower() == ".png"],
        key=_image_sort_key,
    )
    if len(images) <= 1:
        return []
    return list(reversed(images[:-1]))


def build_export_row_from_text(text: str) -> ExportComplaintData:
    transaction_id = _extract_transaction_id(text)
    merchant_order_id = _extract_merchant_order_id(text)
    transaction_amount = _extract_transaction_amount(text)
    transaction_time = _extract_transaction_time(text)
    inquiry_time = _extract_inquiry_time(text)
    problem_description = _extract_problem_description(text)
    refund_summary, refund_status = _detect_refund_status(text)
    return ExportComplaintData(
        transaction_id=transaction_id,
        merchant_order_id=merchant_order_id,
        transaction_amount=transaction_amount,
        transaction_time=transaction_time,
        inquiry_time=inquiry_time,
        problem_description=problem_description,
        refund_summary=refund_summary,
        refund_status=refund_status,
    )


def collect_export_rows_from_run(run_dir: Path, ocr_engine: RapidOCR | None = None) -> list[ExportComplaintData]:
    engine = ocr_engine or RapidOCR()
    rows: list[ExportComplaintData] = []
    complaint_dirs = sorted(
        [path for path in run_dir.iterdir() if path.is_dir()],
        key=_complaint_dir_sort_key,
    )
    for complaint_dir in complaint_dirs:
        text = _ocr_directory_text(complaint_dir, engine)
        row = build_export_row_from_text(text)
        focus_text = _ocr_focus_text_from_directory(complaint_dir, engine)
        if focus_text:
            focused_row = build_export_row_from_text(focus_text)
            if _is_better_transaction_id(focused_row.transaction_id, row.transaction_id):
                row.transaction_id = focused_row.transaction_id
            if _is_better_merchant_order_id(focused_row.merchant_order_id, row.merchant_order_id):
                row.merchant_order_id = focused_row.merchant_order_id
            if _is_preferable_text_value(focused_row.transaction_amount, row.transaction_amount):
                row.transaction_amount = focused_row.transaction_amount
            if _is_preferable_text_value(focused_row.transaction_time, row.transaction_time):
                row.transaction_time = focused_row.transaction_time
            if _is_preferable_text_value(focused_row.inquiry_time, row.inquiry_time):
                row.inquiry_time = focused_row.inquiry_time
            if _is_preferable_text_value(focused_row.problem_description, row.problem_description):
                row.problem_description = focused_row.problem_description
        rows.append(replace(row, complaint_dir=complaint_dir))
    return rows


def export_latest_run_to_workbook(
    run_dir: Path,
    template_path: Path,
    destination_dir: Path,
    rows: list[ExportComplaintData],
) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_path = destination_dir / f"{Path(run_dir).name}_{BASE_OUTPUT_NAME}"

    max_record_images = max((len(list_record_images_for_excel(row.complaint_dir)) for row in rows if row.complaint_dir), default=0)
    image_column_count = max(max_record_images, 1)

    _write_workbook_body(template_path=Path(template_path), output_path=output_path, rows=rows, image_column_count=image_column_count)
    _rewrite_as_wps_cell_image_workbook(output_path=output_path, rows=rows, image_column_count=image_column_count)
    return output_path


def export_latest_run_with_fixed_template(
    base_dir: Path,
    output_root: Path,
    ocr_engine: RapidOCR | None = None,
) -> Path:
    root = Path(output_root)
    if not root.is_absolute():
        root = Path(base_dir) / root

    template_path = Path(base_dir) / "work" / BASE_OUTPUT_NAME
    run_dir = find_latest_run_dir(root)
    rows = collect_export_rows_from_run(run_dir, ocr_engine=ocr_engine)
    return export_latest_run_to_workbook(
        run_dir=run_dir,
        template_path=template_path,
        destination_dir=run_dir,
        rows=rows,
    )


def _write_workbook_body(
    template_path: Path,
    output_path: Path,
    rows: list[ExportComplaintData],
    image_column_count: int,
) -> None:
    workbook = _load_or_create_workbook(template_path)
    sheet = workbook.active
    sheet.title = sheet.title or "Sheet1"

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    _write_headers(sheet, image_column_count)
    _apply_column_widths(sheet, image_column_count)

    for index, row in enumerate(rows, start=2):
        sheet.cell(index, 1, row.transaction_id)
        sheet.cell(index, 2, row.merchant_order_id)
        sheet.cell(index, 3, row.transaction_amount)
        sheet.cell(index, 4, row.transaction_time)
        sheet.cell(index, 5, row.inquiry_time)
        sheet.cell(index, 6, row.problem_description)
        sheet.cell(index, 7, row.refund_summary)
        sheet.cell(index, 8, None)
        sheet.cell(index, 9, None)
        sheet.cell(index, 10, row.refund_status)
        sheet.cell(index, 11, None)
        sheet.cell(index, 12, None)
        sheet.row_dimensions[index].height = DEFAULT_ROW_HEIGHT

    workbook.save(output_path)


def _rewrite_as_wps_cell_image_workbook(
    output_path: Path,
    rows: list[ExportComplaintData],
    image_column_count: int,
) -> None:
    with tempfile.TemporaryDirectory(prefix="wechat-export-") as temp_dir:
        temp_root = Path(temp_dir)
        with ZipFile(output_path) as zf:
            zf.extractall(temp_root)

        xl_dir = temp_root / "xl"
        media_dir = xl_dir / "media"
        if media_dir.exists():
            shutil.rmtree(media_dir)
        media_dir.mkdir(parents=True, exist_ok=True)

        entries = _copy_media_and_build_entries(rows=rows, media_dir=media_dir)
        _update_content_types(temp_root / "[Content_Types].xml")
        _update_workbook_relationships(xl_dir / "_rels" / "workbook.xml.rels")
        _write_cellimages_parts(xl_dir=xl_dir, entries=entries)
        _update_sheet_xml(xl_dir / "worksheets" / "sheet1.xml", rows=rows, entries=entries, image_column_count=image_column_count)

        temp_output = output_path.with_suffix(".tmp.xlsx")
        _zip_directory(temp_root, temp_output)
        temp_output.replace(output_path)


def _load_or_create_workbook(template_path: Path) -> Workbook:
    if template_path.exists():
        return load_workbook(template_path)
    workbook = Workbook()
    return workbook


def _write_headers(sheet, image_column_count: int) -> None:
    for column_index, label in enumerate(HEADER_LABELS, start=1):
        sheet.cell(1, column_index, label)
    for offset in range(image_column_count):
        sheet.cell(1, 13 + offset, RECORD_HEADER)


def _apply_column_widths(sheet, image_column_count: int) -> None:
    base_widths = {
        "A": 32.11,
        "B": 18.67,
        "C": 10.0,
        "D": 20.0,
        "E": 20.0,
        "F": 30.0,
        "G": 8.33,
        "H": 20.89,
        "I": 19.78,
        "J": 20.33,
        "K": 24.11,
        "L": 18.33,
    }
    for column, width in base_widths.items():
        sheet.column_dimensions[column].width = width
    for offset in range(image_column_count):
        sheet.column_dimensions[get_column_letter(13 + offset)].width = 10.0


def _copy_media_and_build_entries(rows: list[ExportComplaintData], media_dir: Path) -> list[_CellImageEntry]:
    entries: list[_CellImageEntry] = []
    media_index = 1
    for row_index, row in enumerate(rows, start=2):
        if row.complaint_dir is None:
            continue
        record_images = list_record_images_for_excel(row.complaint_dir)
        for column_offset, image_path in enumerate(record_images):
            cell_ref = f"{get_column_letter(13 + column_offset)}{row_index}"
            media_name = f"image{media_index}.png"
            shutil.copyfile(image_path, media_dir / media_name)
            entries.append(
                _CellImageEntry(
                    cell_ref=cell_ref,
                    display_id=f"ID_{uuid.uuid4().hex.upper()}",
                    rel_id=f"rId{media_index}",
                    media_name=media_name,
                    description=image_path.stem,
                )
            )
            media_index += 1
    return entries


def _update_content_types(content_types_path: Path) -> None:
    tree = ET.parse(content_types_path)
    root = tree.getroot()

    defaults = root.findall(f"{{{CONTENT_TYPES_NS}}}Default")
    if not any(node.get("Extension") == "png" for node in defaults):
        root.insert(
            0,
            ET.Element(
                f"{{{CONTENT_TYPES_NS}}}Default",
                {"Extension": "png", "ContentType": "image/png"},
            ),
        )

    overrides = root.findall(f"{{{CONTENT_TYPES_NS}}}Override")
    if not any(node.get("PartName") == "/xl/cellimages.xml" for node in overrides):
        root.append(
            ET.Element(
                f"{{{CONTENT_TYPES_NS}}}Override",
                {
                    "PartName": "/xl/cellimages.xml",
                    "ContentType": CELL_IMAGE_CONTENT_TYPE,
                },
            )
        )

    tree.write(content_types_path, encoding="utf-8", xml_declaration=True)


def _update_workbook_relationships(rels_path: Path) -> None:
    tree = ET.parse(rels_path)
    root = tree.getroot()
    relationships = root.findall(f"{{{REL_NS}}}Relationship")
    if any(node.get("Type") == CELL_IMAGE_REL_TYPE for node in relationships):
        tree.write(rels_path, encoding="utf-8", xml_declaration=True)
        return

    next_id = _next_relationship_id(node.get("Id", "") for node in relationships)
    root.append(
        ET.Element(
            f"{{{REL_NS}}}Relationship",
            {
                "Id": next_id,
                "Type": CELL_IMAGE_REL_TYPE,
                "Target": "cellimages.xml",
            },
        )
    )
    tree.write(rels_path, encoding="utf-8", xml_declaration=True)


def _write_cellimages_parts(xl_dir: Path, entries: list[_CellImageEntry]) -> None:
    cellimages_root = ET.Element(f"{{{ETC_NS}}}cellImages")
    for index, entry in enumerate(entries, start=1):
        cell_image = ET.SubElement(cellimages_root, f"{{{ETC_NS}}}cellImage")
        pic = ET.SubElement(cell_image, f"{{{XDR_NS}}}pic")
        nv_pic_pr = ET.SubElement(pic, f"{{{XDR_NS}}}nvPicPr")
        ET.SubElement(
            nv_pic_pr,
            f"{{{XDR_NS}}}cNvPr",
            {
                "id": str(index),
                "name": entry.display_id,
                "descr": entry.description,
            },
        )
        ET.SubElement(nv_pic_pr, f"{{{XDR_NS}}}cNvPicPr")
        blip_fill = ET.SubElement(pic, f"{{{XDR_NS}}}blipFill")
        ET.SubElement(blip_fill, f"{{{A_NS}}}blip", {f"{{{R_NS}}}embed": entry.rel_id})
        stretch = ET.SubElement(blip_fill, f"{{{A_NS}}}stretch")
        ET.SubElement(stretch, f"{{{A_NS}}}fillRect")
        sp_pr = ET.SubElement(pic, f"{{{XDR_NS}}}spPr")
        xfrm = ET.SubElement(sp_pr, f"{{{A_NS}}}xfrm")
        ET.SubElement(xfrm, f"{{{A_NS}}}off", {"x": "0", "y": "0"})
        ET.SubElement(xfrm, f"{{{A_NS}}}ext", {"cx": IMAGE_EXT_CX, "cy": IMAGE_EXT_CY})
        prst_geom = ET.SubElement(sp_pr, f"{{{A_NS}}}prstGeom", {"prst": "rect"})
        ET.SubElement(prst_geom, f"{{{A_NS}}}avLst")

    cellimages_tree = ET.ElementTree(cellimages_root)
    cellimages_tree.write(xl_dir / "cellimages.xml", encoding="utf-8", xml_declaration=True)

    rels_root = ET.Element(f"{{{REL_NS}}}Relationships")
    for entry in entries:
        ET.SubElement(
            rels_root,
            f"{{{REL_NS}}}Relationship",
            {
                "Id": entry.rel_id,
                "Type": IMAGE_REL_TYPE,
                "Target": f"media/{entry.media_name}",
            },
        )
    rels_dir = xl_dir / "_rels"
    rels_dir.mkdir(parents=True, exist_ok=True)
    rels_tree = ET.ElementTree(rels_root)
    rels_tree.write(rels_dir / "cellimages.xml.rels", encoding="utf-8", xml_declaration=True)


def _update_sheet_xml(
    sheet_xml_path: Path,
    rows: list[ExportComplaintData],
    entries: list[_CellImageEntry],
    image_column_count: int,
) -> None:
    tree = ET.parse(sheet_xml_path)
    root = tree.getroot()
    sheet_data = root.find(f"{{{SHEET_NS}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{SHEET_NS}}}sheetData")

    row_map = {int(row.get("r", "0")): row for row in sheet_data.findall(f"{{{SHEET_NS}}}row")}
    entry_map = {entry.cell_ref: entry for entry in entries}

    for row_index in range(2, len(rows) + 2):
        row_node = row_map.get(row_index)
        if row_node is None:
            row_node = ET.SubElement(sheet_data, f"{{{SHEET_NS}}}row", {"r": str(row_index)})
            row_map[row_index] = row_node
        row_node.set("ht", f"{DEFAULT_ROW_HEIGHT}")
        row_node.set("customHeight", "1")

        existing_cells = row_node.findall(f"{{{SHEET_NS}}}c")
        cell_by_ref = {cell.get("r", ""): cell for cell in existing_cells}
        for column_offset in range(image_column_count):
            cell_ref = f"{get_column_letter(13 + column_offset)}{row_index}"
            cell = cell_by_ref.get(cell_ref)
            if cell is None:
                cell = ET.SubElement(row_node, f"{{{SHEET_NS}}}c", {"r": cell_ref})
            else:
                cell.clear()
                cell.set("r", cell_ref)

            entry = entry_map.get(cell_ref)
            if entry is None:
                if cell in list(row_node):
                    row_node.remove(cell)
                continue

            cell.set("t", "str")
            formula = ET.SubElement(cell, f"{{{SHEET_NS}}}f")
            formula.text = f'_xlfn.DISPIMG("{entry.display_id}",1)'
            value = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
            value.text = f'=DISPIMG("{entry.display_id}",1)'

        _sort_row_cells(row_node)

    _sort_sheet_rows(sheet_data)

    last_column_index = 12 + image_column_count
    last_row_index = max(len(rows) + 1, 1)
    dimension = root.find(f"{{{SHEET_NS}}}dimension")
    if dimension is None:
        dimension = ET.Element(f"{{{SHEET_NS}}}dimension")
        root.insert(1, dimension)
    dimension.set("ref", f"A1:{get_column_letter(last_column_index)}{last_row_index}")

    tree.write(sheet_xml_path, encoding="utf-8", xml_declaration=True)


def _sort_row_cells(row_node: ET.Element) -> None:
    cells = row_node.findall(f"{{{SHEET_NS}}}c")
    if not cells:
        return
    for cell in cells:
        row_node.remove(cell)
    for cell in sorted(cells, key=lambda node: _cell_ref_sort_key(node.get("r", ""))):
        row_node.append(cell)


def _sort_sheet_rows(sheet_data: ET.Element) -> None:
    rows = sheet_data.findall(f"{{{SHEET_NS}}}row")
    for row in rows:
        sheet_data.remove(row)
    for row in sorted(rows, key=lambda node: int(node.get("r", "0"))):
        sheet_data.append(row)


def _next_relationship_id(existing_ids: Iterable[str]) -> str:
    max_value = 0
    for rel_id in existing_ids:
        match = re.fullmatch(r"rId(\d+)", rel_id)
        if match:
            max_value = max(max_value, int(match.group(1)))
    return f"rId{max_value + 1}"


def _cell_ref_sort_key(cell_ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", cell_ref)
    if not match:
        return (0, 0)
    column_label, row_number = match.groups()
    column_number = 0
    for char in column_label:
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return (int(row_number), column_number)


def _image_sort_key(path: Path) -> tuple[int, str]:
    match = re.match(r"(\d+)", path.stem)
    if match:
        return (int(match.group(1)), path.name)
    return (10**9, path.name)


def _complaint_dir_sort_key(path: Path) -> tuple[int, str]:
    match = re.match(r"(\d+)", path.name)
    if match:
        return (int(match.group(1)), path.name)
    return (10**9, path.name)


def _ocr_directory_text(complaint_dir: Path, engine: RapidOCR) -> str:
    parts: list[str] = []
    for image_path in sorted(
        [path for path in complaint_dir.iterdir() if path.is_file() and path.suffix.lower() == ".png"],
        key=_image_sort_key,
    ):
        parts.append(_ocr_image_text(image_path, engine))
    return "\n".join(part for part in parts if part)


def _ocr_image_text(image_path: Path, engine: RapidOCR) -> str:
    image = Image.open(image_path)
    try:
        result, _ = engine(np.asarray(image))
    finally:
        image.close()
    if not result:
        return ""
    return "\n".join(text for _box, text, _score in result)


def _ocr_focus_text_from_directory(complaint_dir: Path, engine: RapidOCR) -> str:
    image_paths = sorted(
        [path for path in complaint_dir.iterdir() if path.is_file() and path.suffix.lower() == ".png"],
        key=_image_sort_key,
    )
    if not image_paths:
        return ""

    focus_parts: list[str] = []
    for image_path in image_paths[-1:]:
        focus_parts.append(_ocr_focus_text_from_image(image_path, engine))
    return "\n".join(part for part in focus_parts if part)


def _ocr_focus_text_from_image(image_path: Path, engine: RapidOCR) -> str:
    image = Image.open(image_path)
    try:
        width, height = image.size
        crop = image.crop((0, 0, width, int(height * 0.55)))
        enlarged = crop.resize((crop.width * 2, crop.height * 2))
        result, _ = engine(np.asarray(enlarged))
    finally:
        image.close()
    if not result:
        return ""
    return "\n".join(text for _box, text, _score in result)


def _extract_transaction_id(text: str) -> str:
    direct = _extract_labeled_digits(text, ["交易单号", "交易订单号", "交易流水号"])
    if direct:
        return direct
    candidates = re.findall(r"\d{20,32}", text)
    for candidate in candidates:
        if candidate.startswith("42"):
            return candidate
    return candidates[0] if candidates else ""


def _extract_merchant_order_id(text: str) -> str:
    direct = _extract_labeled_digits(text, ["商户单号", "商户订单号", "商家单号"])
    if direct:
        return direct
    candidates = [candidate for candidate in re.findall(r"\d{8,32}", text) if not candidate.startswith("42")]
    phone_numbers = set(re.findall(r"1\d{10}", text))
    candidates = [candidate for candidate in candidates if candidate not in phone_numbers]
    return candidates[0] if candidates else ""


def _is_better_transaction_id(candidate: str, current: str) -> bool:
    if not candidate:
        return False
    if not current:
        return candidate.startswith("42") and len(candidate) >= 20
    return candidate.startswith("42") and len(candidate) >= len(current)


def _is_better_merchant_order_id(candidate: str, current: str) -> bool:
    if not candidate:
        return False
    if candidate == "4006211990":
        return False
    if not current:
        return len(candidate) >= 12
    if current == "4006211990":
        return len(candidate) >= 12
    return len(candidate) >= len(current)


def _is_preferable_text_value(candidate: str, current: str) -> bool:
    if not candidate:
        return False
    if not current:
        return True
    return len(candidate) >= len(current)


def _extract_transaction_amount(text: str) -> str:
    patterns = [
        r"(?:交易金额|订单金额|支付金额)[^\d¥￥]*[¥￥]?\s*([\d.]+)",
        r"[¥￥]\s*([\d.]+)",
        r"([\d.]+)\s*元",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def _extract_transaction_time(text: str) -> str:
    patterns = [
        r"(?:交易时间|支付时间|下单时间)[^\d]*?(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?\s*\d{1,2}:\d{2}(?::\d{2})?)",
        r"(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            time_str = match.group(1)
            time_str = time_str.replace("年", "-").replace("月", "-").replace("日", "")
            return time_str
    return ""


def _extract_inquiry_time(text: str) -> str:
    patterns = [
        r"(?:咨询时间|投诉时间)[^\d]*?(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?\s*\d{1,2}:\d{2}(?::\d{2})?)",
        r"咨询时间[^\d]*?(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            time_str = match.group(1)
            time_str = time_str.replace("年", "-").replace("月", "-").replace("日", "")
            return time_str
    return ""


def _extract_problem_description(text: str) -> str:
    patterns = [
        r"(?:问题描述|投诉原因|投诉内容)[：:\s]+\s*([^\n]{2,200})",
        r"用户(?:反馈|投诉|表示)[：:\s]+\s*([^\n]{2,200})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            description = match.group(1).strip()
            if len(description) > 100:
                description = description[:100] + "..."
            return description

    lines = text.split("\n")
    for i, line in enumerate(lines):
        if any(keyword in line for keyword in ["问题描述", "投诉原因", "投诉内容"]):
            parts = re.split(r"[：:\s]+", line, maxsplit=1)
            if len(parts) > 1 and parts[1].strip():
                description = parts[1].strip()
                if len(description) > 100:
                    description = description[:100] + "..."
                return description
            if i + 1 < len(lines):
                description = lines[i + 1].strip()
                if len(description) > 100:
                    description = description[:100] + "..."
                return description
        if "咨询原因" in line:
            parts = re.split(r"[：:\s]+", line, maxsplit=1)
            if len(parts) > 1 and parts[1].strip() and parts[1].strip() != "咨询原因":
                description = parts[1].strip()
                if len(description) > 100:
                    description = description[:100] + "..."
                return description
            if i + 1 < len(lines):
                description = lines[i + 1].strip()
                if len(description) > 100:
                    description = description[:100] + "..."
                return description
    return ""


def _extract_labeled_digits(text: str, labels: list[str]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    match = re.search(rf"(?:{label_pattern})[^0-9]{{0,12}}((?:\d[\s]*){{8,40}})", text)
    if not match:
        squashed = re.sub(r"\s+", "", text)
        match = re.search(rf"(?:{label_pattern})[^0-9]{{0,6}}(\d{{8,40}})", squashed)
        if not match:
            return ""
    digits = re.sub(r"\D", "", match.group(1))
    return digits


def _detect_refund_status(text: str) -> tuple[str, str]:
    normalized = re.sub(r"\s+", "", text)
    full_refund_keywords = [
        "全额退款",
        "原路返回",
        "已退款至用户账户",
        "已发起退款",
        "退款成功",
    ]
    if any(keyword in normalized for keyword in full_refund_keywords):
        return ("全额退款", "全额退款（原路返回）")
    return ("未退款", "未退款")


def _zip_directory(source_dir: Path, destination_zip: Path) -> None:
    with ZipFile(destination_zip, "w", compression=ZIP_DEFLATED) as zf:
        for path in sorted(source_dir.rglob("*")):
            if path.is_file():
                zf.write(path, path.relative_to(source_dir).as_posix())
